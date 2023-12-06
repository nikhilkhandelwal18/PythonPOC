import openpyxl as ox

dataframe = ox.load_workbook(r'C:\Test\CPDBA_USER_GROUP_tab.xlsx')

dataframe1 = dataframe.active

# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)